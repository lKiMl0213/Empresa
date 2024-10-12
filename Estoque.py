import openpyxl
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_cors import CORS

app = Flask(__name__)
app.secret_key = 'your_secret_key'

def get_positive_float_input(message):
    while True:
        try:
            value = float(input(message).replace(",", "."))
            if value < 0:
                raise ValueError
            return value
        except ValueError:
            print("Erro: Insira um número positivo válido.")

def get_positive_int_input(message):
    while True:
        try:
            value = int(input(message))
            if value < 0:
                raise ValueError
            return value
        except ValueError:
            print("Erro: Insira um número inteiro positivo válido.")

def get_valid_date_input(message):
    while True:
        expiration_date = input(message).replace("-", "/")
        if len(expiration_date) == 5 and expiration_date[2] == '/':
            try:
                month, year = map(int, expiration_date.split('/'))
                if 1 <= month <= 12:
                    return expiration_date
                else:
                    print("Erro: Mês inválido. O mês deve estar entre 01 e 12.")
            except ValueError:
                print("Erro: Formato de data inválido. Use mm/aa ou mm-aa.")
        else:
            print("Erro: Formato de data inválido. Use mm/aa ou mm-aa.")

def add_storage():
    barcode = input("Digite o código de barras do produto: ")
    workbook = openpyxl.load_workbook("storage.xlsx")
    sheet = workbook.active
    product_found = False
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        if row[0].value == barcode:
            product_found = True
            update = input(f"Produto já cadastrado! Deseja atualizar o produto? (S/N): ").upper()
            if update == "S":
                print("O que deseja atualizar?")
                print("1. Nome")
                print("2. Preço de compra")
                print("3. Preço de venda")
                print("4. Estoque")
                print("5. Data de validade")
                print("6. Voltar ao menu principal")
                choice = input("Escolha uma opção: ")
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    if row[0].value == barcode:
                        if choice == "1":
                            name = input("Digite o novo nome do produto: ")
                            row[1].value = name
                        elif choice == "2":
                            buy_price = get_positive_float_input("Digite o novo preço de compra do produto: ")
                            row[2].value = buy_price
                        elif choice == "3":
                            sell_price = get_positive_float_input("Digite o preço de venda do produto: ")
                            row[3].value = sell_price
                        elif choice == "4":
                            stock = get_positive_int_input("Digite a quantidade que deseja adicionar ao estoque do produto: ")
                            old_stock = row[4].value
                            row[4].value = old_stock + stock
                            print(f"Estoque atualizado com sucesso! Estoque: {row[4].value}")
                        elif choice == "5":
                            expiration_date = get_valid_date_input("Digite a nova data de validade do produto (mm/aa)(mm-aa): ")
                            row[5].value = expiration_date
                        elif choice == "6":
                            print("Voltando ao menu principal...")
                            workbook.close()
                            return
                        else:
                            print("Opção inválida!")
                        break
                save_storage(workbook)
            elif update == "N":
                print("Voltando ao menu principal...")
                workbook.close()
                return
            else:
                print("Opção inválida!")
                workbook.close()
                return
    
    if not product_found:
        name = input("Digite o nome do produto: ")
        buy_price = get_positive_float_input("Digite o preço de compra do produto: ")
        sell_price = get_positive_float_input("Digite o preço de venda do produto: ")
        stock = get_positive_int_input("Digite a quantidade em estoque do produto: ")
        expiration_date = get_valid_date_input("Digite a data de validade do produto (mm/aa)(mm-aa): ")
        new_row = [barcode, name, buy_price, sell_price, stock, expiration_date]
        sheet.append(new_row)
        save_storage(workbook)
    
def save_storage(workbook):
    workbook.save("storage.xlsx")
    workbook.close()
    print("Produto atualizado/cadastrado com sucesso!")
    

def remove_storage():
    barcode = input("Digite o código do produto: ")
    workbook = openpyxl.load_workbook("storage.xlsx")
    sheet = workbook.active
    product_found = False

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        if row[0].value == barcode:
            sheet.delete_rows(row[0].row)
            save_storage(workbook)
            print("Produto removido com sucesso!")
            product_found = True
            break

    if not product_found:
        print("Produto não encontrado!")
        workbook.close()

def low_stock():
    workbook = openpyxl.load_workbook("storage.xlsx")
    sheet = workbook.active
    low_stock_products = []

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
        stock_quantity = row[4].value
        product_name = row[1].value

        if stock_quantity < 5:
            low_stock_products.append((product_name, stock_quantity))

    if low_stock_products:
        print("Produtos com estoque baixo:")
        for product_name, quantity in low_stock_products:
            print(f"{product_name} - Estoque: {quantity}")
    else:
        print("Não há produtos com estoque baixo.")
    workbook.close()
    
def storage_verify():
    try:
        openpyxl.load_workbook("storage.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Código", "Nome", "Preço de Compra", "Preço de Venda", "Estoque", "Data de Validade"])
        save_storage(workbook)

def expiration_date_verify():
    try:
        workbook = openpyxl.load_workbook("storage.xlsx")
        sheet = workbook.active
        produtos_proximos_vencer = []

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            expiration_date = row[5].value
            if expiration_date:
                try:
                    expiration_date = datetime.strptime(expiration_date, "%m/%y")
                    if expiration_date <= datetime.now() + timedelta(days=30):
                        produtos_proximos_vencer.append(row[1].value)
                except ValueError:
                    print(f"Erro ao analisar a data de validade do produto: {row[1].value}")

        if produtos_proximos_vencer:
            print("Produto(s) próximo(s) de vencer:")
            for produto in produtos_proximos_vencer:
                print(produto)
        else:
            print("Não há produtos próximos de vencer.")

    except FileNotFoundError:
        print("Não há produtos cadastrados!")
    except Exception as e:
        print(f"Erro ao verificar produtos próximos de vencer: {e}")
    workbook.close()

def storage_list():
    try:
        workbook = openpyxl.load_workbook("storage.xlsx")
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            print(f"Código de Barras: {row[0]}, Nome do Produto: {row[1]}, "
                  f"Valor de Compra: {row[2]}, Valor de Venda: {row[3]}, "
                  f"Estoque: {row[4]}, Data de Validade: {row[5]}")

    except FileNotFoundError:
        print("Não há produtos cadastrados!")
    workbook.close()

def main():
    while True:
        print("1. Adicionar produto")
        print("2. Remover produto")
        print("3. Listar produtos cadastrados")
        print("4. Sair")
        expiration_date_verify()
        low_stock()
        choice = input("Escolha uma opção: ")

        if choice == "1":
            add_storage()
            storage_verify()
        elif choice == "2":
            remove_storage()
            storage_verify()
        elif choice == "3":
            storage_list()
        elif choice == "4":
            break
        else:
            print("Opção inválida!")

if __name__ == "__main__":
    app.run(debug=True)
