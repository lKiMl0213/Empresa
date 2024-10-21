import openpyxl
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_cors import CORS


app = Flask(__name__)
app.secret_key = 'your_secret_key'


def load_workbook_safe(storage_xlsx):
    try:
        return openpyxl.load_workbook(storage_xlsx)
    except FileNotFoundError:
        print(f"Erro: O arquivo '{storage_xlsx}' não foi encontrado.")
        return None
    except Exception as e:
        print(f"Erro ao carregar o arquivo '{storage_xlsx}': {e}")
        return None

def product_to_cart(cart, subtotal, storage_xlsx):
    barcode = input("Digite o código do produto: ")
    quantity = get_positive_int_input("Digite a quantidade do produto: ")

    workbook = load_workbook_safe(storage_xlsx)
    if workbook is None:
        return cart, subtotal
    
    sheet = workbook.active


    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
        if row[0].value == barcode:
            product_name = row[1].value
            product_price = row[3].value
            stock_quantity = row[4].value
            
            if quantity <= stock_quantity:
                total_price = product_price * quantity
                cart.append({
                    "codigo": barcode,
                    "nome": product_name,
                    "valor": product_price,
                    "quantidade": quantity,
                    "total": total_price
                })
                subtotal += total_price

                print(f"Produto adicionado ao carrinho: {product_name} - Código: {barcode} - Valor: R${product_price:.2f} - Total: R${total_price:.2f}")
                return cart, subtotal
            else:
                print("Quantidade insuficiente em estoque.")
                break
    else:
        print("Produto não encontrado.")
    return cart, subtotal


def remove_product_from_cart(cart, subtotal):
    barcode = input("Digite o código do produto: ")
    quantity = get_positive_int_input("Digite a quantidade que deseja remover do produto: ")
    for item in cart:
        if item["codigo"] == barcode:
            subtotal -= item["total"]
            cart.remove(item)
            print(f"Produto '{item['nome']}' removido do carrinho com sucesso.")
            break
    else:
        print(f"Erro: Produto com código '{barcode}' não encontrado no carrinho.")
    return cart, subtotal


def finish_purchase(cart, subtotal):
    print("Compra finalizada.")
    print("Carrinho de Compras:")
    for item in cart:
        print(f"Código: {item['codigo']}, Nome: {item['nome']}, Valor: R${item['valor']:.2f}, Quantidade: {item['quantidade']}, Total: R${item['total']:.2f}")
    print(f"\nSubtotal: R${subtotal:.2f}")


def price_verification():
    barcode = input("Digite o código do produto: ")
    workbook = openpyxl.load_workbook("storage.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
        if row[0].value == barcode:
            product_name = row[1].value
            product_price = row[3].value
            print(f"O valor do produto {product_name} é R${product_price:.2f}")
            break
    else:
        print("Produto não encontrado.")


def update_stock(cart):
    workbook = openpyxl.load_workbook("storage.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
        for item in cart:
            if row[0].value == item["codigo"]:
                row[4].value -= item["quantidade"]
    
    save_storage(workbook)

def save_storage(workbook):
    try:
        workbook.save("storage.xlsx")
        print("Dados salvos com sucesso!")
    except Exception as e:
        print(f"Erro ao salvar o arquivo: {e}")
    finally:
        workbook.close()


def finish_purchase_menu(cart, subtotal):
    print("Carrinho de Compras:")
    for item in cart:
        print(f"Código: {item['codigo']}, Nome: {item['nome']}, Valor: R${item['valor']:.2f}, Quantidade: {item['quantidade']}, Total: R${item['total']:.2f}")
    
    print(f"\nSubtotal: R${subtotal:.2f}")
    payment_method = input("Digite o método de pagamento (dinheiro(D)/cartão(C)), para voltar ao menu digite (V): ").upper()
    if payment_method == 'D' or payment_method == 'C':
        print(f"Compra finalizada. Total a pagar: R${subtotal:.2f}")
        finish_purchase(cart, subtotal)
        update_stock(cart)
        cart.clear()
        subtotal = 0.0
    elif payment_method == 'V':
        return cart, subtotal
    else:
        print("Método de pagamento inválido.")
    
    return cart, subtotal


def main(cart, subtotal):
    print("Bem-vindo ao mercado!")
    print("1 - Adicionar produto ao carrinho")
    print("2 - Finalizar compra")
    print("3 - Remover produto do carrinho")
    print("4 - Verificar preço de um produto")
    print("5 - Sair")

    while True:
        option = input("Digite a opção desejada: ")
        if option == "1":
            cart, subtotal = product_to_cart(cart, subtotal)
        elif option == "2":
            cart, subtotal = finish_purchase_menu(cart, subtotal)
        elif option == "3":
            cart, subtotal = remove_product_from_cart(cart, subtotal)
        elif option == "4":
            price_verification()
        elif option == "5":
            break
        else:
            print("Opção inválida.")


if __name__ == "__main__":
    app.run(debug=True)
